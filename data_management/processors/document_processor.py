"""📄 Document Processor - IA Influencer Agent Platform Enterprise
===============================================================
Module: backend/data_management/processors/document_processor.py
Author: Fahed Mlaiel (mlaiel@live.de)
Team: Lead Dev IA + Backend Senior + ML Engineer + DBA + Security + Microservices + Audio + DevOps + IA Prompt Engineer
Type: Industrial Document Processing - Enterprise Production-Ready
Responsibility: Traitement avancé des documents pour créateurs de contenu textuel
===================================================================================

⚠️  PROPRIÉTÉ INTELLECTUELLE EXCLUSIVE - FAHED MLAIEL ⚠️
© 2025 Fahed Mlaiel. Tous droits réservés.
Usage non autorisé strictement interdit et passible de poursuites judiciaires.
Contact: mlaiel@live.de

LOGIQUE MÉTIER DOCUMENT PROCESSOR:
Document Upload → Format Detection → Text Extraction → NLP Analysis → 
Content Classification → Sentiment Analysis → SEO Analysis → Protection Fingerprinting
"""
import os
import re
import hashlib
from typing import Dict, List, Optional, Any, Union, Tuple
import asyncio
import aiofiles
from concurrent.futures import ThreadPoolExecutor
import logging
from datetime import datetime, timezone
from pathlib import Path

# Document processing libraries
import PyPDF2
import docx
import openpyxl
from bs4 import BeautifulSoup
import markdown
import textract

# NLP and ML libraries
import spacy
import nltk
from textblob import TextBlob
from transformers import pipeline, AutoTokenizer, AutoModel
import torch
from sentence_transformers import SentenceTransformer
import pandas as pd

# Text analysis
from wordcloud import WordCloud
import matplotlib.pyplot as plt
from collections import Counter
import language_tool_python

from .base_processor import BaseProcessor, AsyncBaseProcessor


class DocumentProcessor(BaseProcessor):
    """Processeur avancé pour documents - Production Enterprise"""
    
    def __init__(self, config: Dict[str, Any] = None):
        super().__init__(config)
        self.supported_formats = {
            'PDF', 'DOCX', 'DOC', 'TXT', 'RTF', 'ODT', 'HTML', 'XML', 
            'MD', 'MARKDOWN', 'CSV', 'XLSX', 'XLS', 'JSON', 'EPUB'
        }
        
        self.max_file_size = 100 * 1024 * 1024  # 100MB
        self.max_text_length = 1000000  # 1M characters
        
        # Initialize NLP models
        self._init_nlp_models()
        
        # Content analysis settings
        self.readability_thresholds = {
            'excellent': 90,
            'good': 70,
            'acceptable': 50,
            'difficult': 30
        }
        
        self.logger = logging.getLogger(__name__)
    
    def _init_nlp_models(self):
        """Initialize NLP models and tools"""
        try:
            # Load spaCy model for advanced NLP
            self.nlp = spacy.load("en_core_web_sm")
            
            # Sentiment analysis pipeline
            self.sentiment_analyzer = pipeline(
                "sentiment-analysis",
                model="cardiffnlp/twitter-roberta-base-sentiment-latest"
            )
            
            # Text classification pipeline
            self.text_classifier = pipeline(
                "zero-shot-classification",
                model="facebook/bart-large-mnli"
            )
            
            # Sentence transformer for embeddings
            self.sentence_model = SentenceTransformer('all-MiniLM-L6-v2')
            
            # Grammar checker
            self.grammar_tool = language_tool_python.LanguageTool('en-US')
            
            # Download required NLTK data
            nltk.download('punkt', quiet=True)
            nltk.download('stopwords', quiet=True)
            nltk.download('vader_lexicon', quiet=True)
            
        except Exception as e:
            self.logger.warning(f"NLP models initialization warning: {e}")
            self.nlp = None
            self.sentiment_analyzer = None
            self.text_classifier = None
            self.sentence_model = None
            self.grammar_tool = None
    
    def validate_input(self, input_data: Any) -> bool:
        """Valide les données document d'entrée"""
        if isinstance(input_data, str):
            # File path validation
            path = Path(input_data)
            return (path.exists() and 
                   path.suffix.upper()[1:] in self.supported_formats and
                   path.stat().st_size <= self.max_file_size)
        elif isinstance(input_data, bytes):
            # Binary data validation
            return 0 < len(input_data) <= self.max_file_size
        elif hasattr(input_data, 'read'):
            # File-like object
            return True
        elif isinstance(input_data, dict) and 'content' in input_data:
            # Structured text data
            return isinstance(input_data['content'], str)
        
        return False
    
    def process(self, input_data: Any) -> Dict[str, Any]:
        """Traite un document complètement"""
        try:
            # Extract text content
            text_content = self._extract_text(input_data)
            
            # Basic document metadata
            metadata = self._extract_metadata(input_data, text_content)
            
            # Text analysis
            text_analysis = self._analyze_text(text_content)
            
            # Content classification
            content_classification = self._classify_content(text_content)
            
            # SEO analysis
            seo_analysis = self._analyze_seo(text_content)
            
            # Security analysis
            security_analysis = self._security_analysis(text_content)
            
            # Generate fingerprints
            fingerprints = self._generate_fingerprints(text_content)
            
            # Readability analysis
            readability = self._analyze_readability(text_content)
            
            # Quality assessment
            quality_assessment = self._assess_quality(text_content, text_analysis)
            
            return {
                "success": True,
                "metadata": metadata,
                "text_analysis": text_analysis,
                "content_classification": content_classification,
                "seo_analysis": seo_analysis,
                "security_analysis": security_analysis,
                "fingerprints": fingerprints,
                "readability": readability,
                "quality_assessment": quality_assessment,
                "processing_info": {
                    "processor_version": "3.0.0",
                    "processed_at": datetime.now(timezone.utc).isoformat(),
                    "text_length": len(text_content),
                    "language": metadata.get("language", "unknown")
                }
            }
            
        except Exception as e:
            self.logger.error(f"Document processing error: {e}")
            return {
                "success": False,
                "error": str(e),
                "error_type": type(e).__name__
            }
    
    def _extract_text(self, input_data: Any) -> str:
        """Extrait le texte du document"""
        if isinstance(input_data, str):
            # File path
            return self._extract_text_from_file(input_data)
        elif isinstance(input_data, bytes):
            # Binary data - try to decode as text
            try:
                return input_data.decode('utf-8')
            except UnicodeDecodeError:
                return input_data.decode('latin-1', errors='ignore')
        elif hasattr(input_data, 'read'):
            # File-like object
            content = input_data.read()
            if isinstance(content, bytes):
                return content.decode('utf-8', errors='ignore')
            return content
        elif isinstance(input_data, dict) and 'content' in input_data:
            # Structured data
            return input_data['content']
        
        return str(input_data)
    
    def _extract_text_from_file(self, file_path: str) -> str:
        """Extrait le texte selon le format de fichier"""
        path = Path(file_path)
        extension = path.suffix.upper()
        
        try:
            if extension == '.PDF':
                return self._extract_pdf_text(file_path)
            elif extension in ['.DOCX', '.DOC']:
                return self._extract_docx_text(file_path)
            elif extension in ['.XLSX', '.XLS']:
                return self._extract_excel_text(file_path)
            elif extension in ['.HTML', '.HTM']:
                return self._extract_html_text(file_path)
            elif extension in ['.MD', '.MARKDOWN']:
                return self._extract_markdown_text(file_path)
            elif extension in ['.TXT', '.RTF']:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    return f.read()
            else:
                # Try textract for other formats
                return textract.process(file_path).decode('utf-8', errors='ignore')
        
        except Exception as e:
            self.logger.warning(f"Text extraction failed for {file_path}: {e}")
            # Fallback to basic file reading
            try:
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                    return f.read()
            except:
                return ""
    
    def _extract_pdf_text(self, file_path: str) -> str:
        """Extrait le texte d'un PDF"""
        text = ""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
        except Exception as e:
            self.logger.warning(f"PDF extraction error: {e}")
        
        return text
    
    def _extract_docx_text(self, file_path: str) -> str:
        """Extrait le texte d'un DOCX"""
        try:
            doc = docx.Document(file_path)
            text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
            return text
        except Exception as e:
            self.logger.warning(f"DOCX extraction error: {e}")
            return ""
    
    def _extract_excel_text(self, file_path: str) -> str:
        """Extrait le texte d'un Excel"""
        try:
            workbook = openpyxl.load_workbook(file_path)
            text = ""
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    row_text = " ".join([str(cell) for cell in row if cell is not None])
                    text += row_text + "\n"
            return text
        except Exception as e:
            self.logger.warning(f"Excel extraction error: {e}")
            return ""
    
    def _extract_html_text(self, file_path: str) -> str:
        """Extrait le texte d'un HTML"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                soup = BeautifulSoup(f.read(), 'html.parser')
                return soup.get_text()
        except Exception as e:
            self.logger.warning(f"HTML extraction error: {e}")
            return ""
    
    def _extract_markdown_text(self, file_path: str) -> str:
        """Extrait le texte d'un Markdown"""
        try:
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                md_content = f.read()
                # Convert to HTML then extract text
                html = markdown.markdown(md_content)
                soup = BeautifulSoup(html, 'html.parser')
                return soup.get_text()
        except Exception as e:
            self.logger.warning(f"Markdown extraction error: {e}")
            return ""
    
    def _extract_metadata(self, input_data: Any, text_content: str) -> Dict[str, Any]:
        """Extrait les métadonnées du document"""
        metadata = {
            "text_length": len(text_content),
            "word_count": len(text_content.split()),
            "character_count": len(text_content),
            "paragraph_count": len([p for p in text_content.split('\n\n') if p.strip()]),
            "line_count": len(text_content.split('\n')),
            "language": self._detect_language(text_content)
        }
        
        # File-specific metadata
        if isinstance(input_data, str):
            path = Path(input_data)
            metadata.update({
                "filename": path.name,
                "file_extension": path.suffix,
                "file_size_bytes": path.stat().st_size,
                "created_at": datetime.fromtimestamp(path.stat().st_ctime).isoformat(),
                "modified_at": datetime.fromtimestamp(path.stat().st_mtime).isoformat()
            })
        
        return metadata
    
    def _analyze_text(self, text_content: str) -> Dict[str, Any]:
        """Analyse complète du texte"""
        analysis = {
            "basic_stats": self._get_basic_text_stats(text_content),
            "linguistic_features": self._analyze_linguistic_features(text_content),
            "sentiment_analysis": self._analyze_sentiment(text_content),
            "named_entities": self._extract_named_entities(text_content),
            "keywords": self._extract_keywords(text_content),
            "topics": self._extract_topics(text_content)
        }
        
        return analysis
    
    def _get_basic_text_stats(self, text: str) -> Dict[str, Any]:
        """Statistiques textuelles de base"""
        words = text.split()
        sentences = re.split(r'[.!?]+', text)
        sentences = [s.strip() for s in sentences if s.strip()]
        
        # Calculate averages
        avg_word_length = sum(len(word) for word in words) / len(words) if words else 0
        avg_sentence_length = len(words) / len(sentences) if sentences else 0
        
        # Count unique words
        unique_words = set(word.lower().strip('.,!?;:"()[]{}') for word in words)
        vocabulary_richness = len(unique_words) / len(words) if words else 0
        
        return {
            "total_words": len(words),
            "total_sentences": len(sentences),
            "unique_words": len(unique_words),
            "average_word_length": round(avg_word_length, 2),
            "average_sentence_length": round(avg_sentence_length, 2),
            "vocabulary_richness": round(vocabulary_richness, 2),
            "most_common_words": Counter(word.lower() for word in words).most_common(10)
        }
    
    def _analyze_linguistic_features(self, text: str) -> Dict[str, Any]:
        """Analyse des caractéristiques linguistiques"""
        if not self.nlp:
            return {"error": "NLP model not available"}
        
        # Process with spaCy
        doc = self.nlp(text[:1000000])  # Limit text length for performance
        
        # Extract linguistic features
        pos_counts = Counter(token.pos_ for token in doc)
        dep_counts = Counter(token.dep_ for token in doc)
        
        return {
            "part_of_speech_distribution": dict(pos_counts),
            "dependency_relations": dict(dep_counts.most_common(10)),
            "sentence_complexity": self._calculate_sentence_complexity(doc),
            "passive_voice_percentage": self._detect_passive_voice(doc),
            "lexical_diversity": len(set(token.lemma_.lower() for token in doc if token.is_alpha)) / len([token for token in doc if token.is_alpha])
        }
    
    def _analyze_sentiment(self, text: str) -> Dict[str, Any]:
        """Analyse du sentiment"""
        sentiment_results = {}
        
        # TextBlob sentiment
        blob = TextBlob(text)
        sentiment_results["textblob"] = {
            "polarity": blob.sentiment.polarity,
            "subjectivity": blob.sentiment.subjectivity
        }
        
        # Transformer-based sentiment if available
        if self.sentiment_analyzer:
            try:
                # Truncate text for transformer model
                truncated_text = text[:512]
                transformer_sentiment = self.sentiment_analyzer(truncated_text)[0]
                sentiment_results["transformer"] = {
                    "label": transformer_sentiment['label'],
                    "score": transformer_sentiment['score']
                }
            except Exception as e:
                self.logger.warning(f"Transformer sentiment analysis failed: {e}")
        
        # VADER sentiment from NLTK
        try:
            from nltk.sentiment import SentimentIntensityAnalyzer
            analyzer = SentimentIntensityAnalyzer()
            vader_scores = analyzer.polarity_scores(text)
            sentiment_results["vader"] = vader_scores
        except Exception as e:
            self.logger.warning(f"VADER sentiment analysis failed: {e}")
        
        return sentiment_results
    
    def _extract_named_entities(self, text: str) -> List[Dict[str, Any]]:
        """Extraction d'entités nommées"""
        if not self.nlp:
            return []
        
        doc = self.nlp(text[:1000000])
        entities = []
        
        for ent in doc.ents:
            entities.append({
                "text": ent.text,
                "label": ent.label_,
                "description": spacy.explain(ent.label_),
                "start": ent.start_char,
                "end": ent.end_char
            })
        
        return entities
    
    def _extract_keywords(self, text: str) -> List[Dict[str, Any]]:
        """Extraction de mots-clés"""
        if not self.nlp:
            return []
        
        doc = self.nlp(text[:1000000])
        
        # Extract significant terms (nouns, adjectives, proper nouns)
        keywords = []
        for token in doc:
            if (token.pos_ in ['NOUN', 'ADJ', 'PROPN'] and 
                not token.is_stop and 
                not token.is_punct and 
                len(token.text) > 2):
                keywords.append(token.lemma_.lower())
        
        # Count and rank keywords
        keyword_counts = Counter(keywords)
        
        return [
            {"keyword": word, "frequency": count, "relevance_score": count / len(keywords)}
            for word, count in keyword_counts.most_common(20)
        ]
    
    def _extract_topics(self, text: str) -> List[Dict[str, Any]]:
        """Extraction de sujets/thèmes"""
        if not self.text_classifier:
            return []
        
        # Predefined topic candidates
        candidate_labels = [
            "technology", "business", "science", "health", "education",
            "entertainment", "sports", "politics", "travel", "food",
            "finance", "marketing", "lifestyle", "news", "opinion"
        ]
        
        try:
            # Truncate text for classification
            truncated_text = text[:1024]
            classification = self.text_classifier(truncated_text, candidate_labels)
            
            topics = []
            for label, score in zip(classification['labels'], classification['scores']):
                topics.append({
                    "topic": label,
                    "confidence": score,
                    "relevance": "high" if score > 0.7 else "medium" if score > 0.4 else "low"
                })
            
            return topics[:5]  # Top 5 topics
            
        except Exception as e:
            self.logger.warning(f"Topic extraction failed: {e}")
            return []
    
    def _classify_content(self, text: str) -> Dict[str, Any]:
        """Classification du contenu"""
        classification = {
            "content_type": self._determine_content_type(text),
            "writing_style": self._analyze_writing_style(text),
            "target_audience": self._determine_target_audience(text),
            "content_purpose": self._determine_content_purpose(text)
        }
        
        return classification
    
    def _analyze_seo(self, text: str) -> Dict[str, Any]:
        """Analyse SEO du contenu"""
        words = text.split()
        
        # Basic SEO metrics
        seo_analysis = {
            "word_count": len(words),
            "recommended_word_count": self._get_seo_word_count_recommendation(text),
            "keyword_density": self._calculate_keyword_density(text),
            "readability_score": self._calculate_readability_score(text),
            "heading_structure": self._analyze_heading_structure(text),
            "content_quality_score": self._calculate_content_quality_score(text),
            "seo_recommendations": self._generate_seo_recommendations(text)
        }
        
        return seo_analysis
    
    def _security_analysis(self, text: str) -> Dict[str, Any]:
        """Analyse de sécurité du contenu"""
        security_issues = []
        
        # Check for sensitive information patterns
        patterns = {
            "email": r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b',
            "phone": r'\b\d{3}[-.]?\d{3}[-.]?\d{4}\b',
            "ssn": r'\b\d{3}-\d{2}-\d{4}\b',
            "credit_card": r'\b\d{4}[-\s]?\d{4}[-\s]?\d{4}[-\s]?\d{4}\b',
            "ip_address": r'\b\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}\b'
        }
        
        for pattern_name, pattern in patterns.items():
            matches = re.findall(pattern, text)
            if matches:
                security_issues.append({
                    "type": pattern_name,
                    "count": len(matches),
                    "risk_level": "high" if pattern_name in ["ssn", "credit_card"] else "medium"
                })
        
        return {
            "security_issues": security_issues,
            "privacy_risk_score": self._calculate_privacy_risk(security_issues),
            "content_safety_score": self._assess_content_safety(text),
            "profanity_detected": self._detect_profanity(text)
        }
    
    def _generate_fingerprints(self, text: str) -> Dict[str, Any]:
        """Génère des empreintes du texte"""
        # Text hashes
        md5_hash = hashlib.md5(text.encode()).hexdigest()
        sha256_hash = hashlib.sha256(text.encode()).hexdigest()
        
        # Semantic fingerprint using sentence transformers
        semantic_fingerprint = []
        if self.sentence_model:
            try:
                # Create semantic embedding
                embedding = self.sentence_model.encode(text[:512])  # Limit text length
                semantic_fingerprint = embedding.tolist()[:64]  # Reduce dimensionality
            except Exception as e:
                self.logger.warning(f"Semantic fingerprint generation failed: {e}")
        
        # N-gram fingerprints
        ngram_fingerprints = self._generate_ngram_fingerprints(text)
        
        return {
            "md5_hash": md5_hash,
            "sha256_hash": sha256_hash,
            "semantic_fingerprint": semantic_fingerprint,
            "ngram_fingerprints": ngram_fingerprints,
            "text_length": len(text),
            "word_count": len(text.split()),
            "fingerprint_version": "2.0"
        }
    
    def _analyze_readability(self, text: str) -> Dict[str, Any]:
        """Analyse de lisibilité"""
        return {
            "flesch_kincaid_grade": self._calculate_flesch_kincaid(text),
            "flesch_reading_ease": self._calculate_flesch_reading_ease(text),
            "automated_readability_index": self._calculate_ari(text),
            "gunning_fog_index": self._calculate_gunning_fog(text),
            "readability_rating": self._get_readability_rating(text)
        }
    
    def _assess_quality(self, text: str, text_analysis: Dict[str, Any]) -> Dict[str, Any]:
        """Évaluation de la qualité du contenu"""
        quality_scores = {
            "grammar_score": self._check_grammar(text),
            "coherence_score": self._assess_coherence(text),
            "completeness_score": self._assess_completeness(text),
            "originality_score": self._assess_originality(text),
            "engagement_score": self._assess_engagement_potential(text)
        }
        
        # Overall quality score
        overall_score = sum(quality_scores.values()) / len(quality_scores)
        
        return {
            "individual_scores": quality_scores,
            "overall_quality_score": overall_score,
            "quality_rating": self._get_quality_rating(overall_score),
            "improvement_suggestions": self._generate_improvement_suggestions(quality_scores, text_analysis)
        }
    
    # Utility methods
    def _detect_language(self, text: str) -> str:
        """Détecte la langue du texte"""
        try:
            from langdetect import detect
            return detect(text)
        except:
            return "unknown"
    
    def _calculate_sentence_complexity(self, doc) -> float:
        """Calcule la complexité des phrases"""
        if not doc.sents:
            return 0.0
        
        complexities = []
        for sent in doc.sents:
            # Count clauses based on conjunctions and punctuation
            clause_markers = len([token for token in sent if token.dep_ in ['ccomp', 'xcomp', 'advcl']])
            complexity = clause_markers / len(sent) if len(sent) > 0 else 0
            complexities.append(complexity)
        
        return sum(complexities) / len(complexities)
    
    def _detect_passive_voice(self, doc) -> float:
        """Détecte le pourcentage de voix passive"""
        passive_count = 0
        total_sentences = 0
        
        for sent in doc.sents:
            total_sentences += 1
            # Look for passive voice patterns
            for token in sent:
                if token.dep_ == 'auxpass' or (token.tag_ == 'VBN' and any(child.dep_ == 'auxpass' for child in token.children)):
                    passive_count += 1
                    break
        
        return (passive_count / total_sentences * 100) if total_sentences > 0 else 0
    
    def _determine_content_type(self, text: str) -> str:
        """Détermine le type de contenu"""
        # Simple heuristics based on text characteristics
        if len(text.split()) < 100:
            return "short_form"
        elif re.search(r'\n\s*[-*+]\s+', text):  # List patterns
            return "list_article"
        elif re.search(r'\b(how to|step \d+|first|second|third)\b', text, re.IGNORECASE):
            return "how_to_guide"
        elif re.search(r'\b(review|rating|stars|recommend)\b', text, re.IGNORECASE):
            return "review"
        elif re.search(r'\b(news|breaking|reported|according to)\b', text, re.IGNORECASE):
            return "news_article"
        else:
            return "general_article"
    
    def _analyze_writing_style(self, text: str) -> Dict[str, Any]:
        """Analyse le style d'écriture"""
        sentences = re.split(r'[.!?]+', text)
        avg_sentence_length = sum(len(s.split()) for s in sentences) / len(sentences) if sentences else 0
        
        # Determine writing style
        if avg_sentence_length < 10:
            style = "conversational"
        elif avg_sentence_length > 25:
            style = "academic"
        else:
            style = "professional"
        
        return {
            "style": style,
            "average_sentence_length": avg_sentence_length,
            "formality_level": self._assess_formality(text),
            "tone": self._assess_tone(text)
        }
    
    def _determine_target_audience(self, text: str) -> str:
        """Détermine l'audience cible"""
        readability_score = self._calculate_flesch_kincaid(text)
        
        if readability_score < 6:
            return "general_public"
        elif readability_score < 10:
            return "educated_general"
        elif readability_score < 13:
            return "college_level"
        else:
            return "expert_professional"
    
    def _determine_content_purpose(self, text: str) -> str:
        """Détermine l'objectif du contenu"""
        # Simple keyword-based classification
        if re.search(r'\b(buy|purchase|order|sale|discount)\b', text, re.IGNORECASE):
            return "commercial"
        elif re.search(r'\b(learn|education|tutorial|guide)\b', text, re.IGNORECASE):
            return "educational"
        elif re.search(r'\b(entertain|funny|humor|story)\b', text, re.IGNORECASE):
            return "entertainment"
        elif re.search(r'\b(inform|news|report|announce)\b', text, re.IGNORECASE):
            return "informational"
        else:
            return "general"
    
    def _get_seo_word_count_recommendation(self, text: str) -> Dict[str, Any]:
        """Recommandations de nombre de mots pour SEO"""
        word_count = len(text.split())
        
        if word_count < 300:
            return {"status": "too_short", "recommendation": "Aim for 300+ words for better SEO"}
        elif word_count < 1000:
            return {"status": "good", "recommendation": "Good length for most content"}
        elif word_count < 2000:
            return {"status": "excellent", "recommendation": "Excellent length for in-depth content"}
        else:
            return {"status": "very_long", "recommendation": "Consider breaking into multiple articles"}
    
    def _calculate_keyword_density(self, text: str) -> Dict[str, Any]:
        """Calcule la densité des mots-clés"""
        words = text.lower().split()
        word_counts = Counter(words)
        total_words = len(words)
        
        # Calculate density for top words
        densities = {}
        for word, count in word_counts.most_common(10):
            if len(word) > 3:  # Ignore short words
                density = (count / total_words) * 100
                densities[word] = round(density, 2)
        
        return densities
    
    def _calculate_readability_score(self, text: str) -> float:
        """Calcule un score de lisibilité composite"""
        flesch_ease = self._calculate_flesch_reading_ease(text)
        return flesch_ease / 100  # Normalize to 0-1
    
    def _analyze_heading_structure(self, text: str) -> Dict[str, Any]:
        """Analyse la structure des titres"""
        # Look for markdown or HTML headings
        h1_count = len(re.findall(r'^#\s+', text, re.MULTILINE))
        h2_count = len(re.findall(r'^##\s+', text, re.MULTILINE))
        h3_count = len(re.findall(r'^###\s+', text, re.MULTILINE))
        
        return {
            "h1_count": h1_count,
            "h2_count": h2_count,
            "h3_count": h3_count,
            "has_proper_structure": h1_count == 1 and h2_count > 0,
            "recommendations": self._generate_heading_recommendations(h1_count, h2_count, h3_count)
        }
    
    def _calculate_content_quality_score(self, text: str) -> float:
        """Calcule un score de qualité du contenu"""
        # Composite score based on multiple factors
        word_count = len(text.split())
        unique_words = len(set(word.lower() for word in text.split()))
        
        # Length score
        length_score = min(word_count / 1000, 1.0) * 0.3
        
        # Vocabulary diversity score
        diversity_score = (unique_words / word_count) * 0.3 if word_count > 0 else 0
        
        # Readability score
        readability_score = self._calculate_readability_score(text) * 0.4
        
        return length_score + diversity_score + readability_score
    
    def _generate_seo_recommendations(self, text: str) -> List[str]:
        """Génère des recommandations SEO"""
        recommendations = []
        
        word_count = len(text.split())
        if word_count < 300:
            recommendations.append("Increase content length to at least 300 words")
        
        # Check for headings
        if not re.search(r'^#+\s+', text, re.MULTILINE):
            recommendations.append("Add headings to improve content structure")
        
        # Check for lists
        if not re.search(r'\n\s*[-*+]\s+', text):
            recommendations.append("Consider adding bullet points or numbered lists")
        
        recommendations.extend([
            "Optimize meta description",
            "Add internal and external links",
            "Include relevant keywords naturally",
            "Add alt text for images",
            "Ensure mobile-friendly formatting"
        ])
        
        return recommendations
    
    def _calculate_privacy_risk(self, security_issues: List[Dict]) -> float:
        """Calcule le risque de confidentialité"""
        if not security_issues:
            return 0.0
        
        risk_weights = {"ssn": 1.0, "credit_card": 1.0, "email": 0.3, "phone": 0.5, "ip_address": 0.2}
        total_risk = sum(issue["count"] * risk_weights.get(issue["type"], 0.1) for issue in security_issues)
        
        return min(total_risk / 10, 1.0)  # Normalize to 0-1
    
    def _assess_content_safety(self, text: str) -> float:
        """Évalue la sécurité du contenu"""
        # Basic content safety assessment
        # In production, this would use specialized content moderation APIs
        return 0.95  # Placeholder
    
    def _detect_profanity(self, text: str) -> bool:
        """Détecte la profanité"""
        # Basic profanity detection
        # In production, use specialized libraries like better-profanity
        return False  # Placeholder
    
    def _generate_ngram_fingerprints(self, text: str) -> Dict[str, List[str]]:
        """Génère des empreintes n-grammes"""
        words = text.lower().split()
        
        # Generate 2-grams and 3-grams
        bigrams = [f"{words[i]}_{words[i+1]}" for i in range(len(words)-1)]
        trigrams = [f"{words[i]}_{words[i+1]}_{words[i+2]}" for i in range(len(words)-2)]
        
        return {
            "bigrams": bigrams[:20],  # Top 20 bigrams
            "trigrams": trigrams[:20]  # Top 20 trigrams
        }
    
    def _calculate_flesch_kincaid(self, text: str) -> float:
        """Calcule le grade Flesch-Kincaid"""
        sentences = re.split(r'[.!?]+', text)
        sentences = [s.strip() for s in sentences if s.strip()]
        
        words = text.split()
        syllables = sum(self._count_syllables(word) for word in words)
        
        if len(sentences) == 0 or len(words) == 0:
            return 0
        
        return 0.39 * (len(words) / len(sentences)) + 11.8 * (syllables / len(words)) - 15.59
    
    def _calculate_flesch_reading_ease(self, text: str) -> float:
        """Calcule la facilité de lecture Flesch"""
        sentences = re.split(r'[.!?]+', text)
        sentences = [s.strip() for s in sentences if s.strip()]
        
        words = text.split()
        syllables = sum(self._count_syllables(word) for word in words)
        
        if len(sentences) == 0 or len(words) == 0:
            return 0
        
        return 206.835 - 1.015 * (len(words) / len(sentences)) - 84.6 * (syllables / len(words))
    
    def _calculate_ari(self, text: str) -> float:
        """Calcule l'Automated Readability Index"""
        sentences = re.split(r'[.!?]+', text)
        sentences = [s.strip() for s in sentences if s.strip()]
        
        words = text.split()
        characters = sum(len(word) for word in words)
        
        if len(sentences) == 0 or len(words) == 0:
            return 0
        
        return 4.71 * (characters / len(words)) + 0.5 * (len(words) / len(sentences)) - 21.43
    
    def _calculate_gunning_fog(self, text: str) -> float:
        """Calcule l'indice Gunning Fog"""
        sentences = re.split(r'[.!?]+', text)
        sentences = [s.strip() for s in sentences if s.strip()]
        
        words = text.split()
        complex_words = [word for word in words if self._count_syllables(word) > 2]
        
        if len(sentences) == 0 or len(words) == 0:
            return 0
        
        return 0.4 * ((len(words) / len(sentences)) + 100 * (len(complex_words) / len(words)))
    
    def _get_readability_rating(self, text: str) -> str:
        """Obtient une évaluation de lisibilité"""
        flesch_ease = self._calculate_flesch_reading_ease(text)
        
        if flesch_ease >= 90:
            return "very_easy"
        elif flesch_ease >= 80:
            return "easy"
        elif flesch_ease >= 70:
            return "fairly_easy"
        elif flesch_ease >= 60:
            return "standard"
        elif flesch_ease >= 50:
            return "fairly_difficult"
        elif flesch_ease >= 30:
            return "difficult"
        else:
            return "very_difficult"
    
    def _count_syllables(self, word: str) -> int:
        """Compte les syllabes dans un mot"""
        word = word.lower()
        vowels = "aeiouy"
        syllable_count = 0
        previous_was_vowel = False
        
        for char in word:
            is_vowel = char in vowels
            if is_vowel and not previous_was_vowel:
                syllable_count += 1
            previous_was_vowel = is_vowel
        
        # Handle silent e
        if word.endswith('e'):
            syllable_count -= 1
        
        # Every word has at least one syllable
        return max(1, syllable_count)
    
    def _check_grammar(self, text: str) -> float:
        """Vérifie la grammaire"""
        if not self.grammar_tool:
            return 0.8  # Default score
        
        try:
            matches = self.grammar_tool.check(text[:5000])  # Limit text length
            error_count = len(matches)
            word_count = len(text.split())
            
            # Calculate grammar score (fewer errors = higher score)
            error_rate = error_count / word_count if word_count > 0 else 0
            grammar_score = max(0, 1 - error_rate * 10)  # Scale appropriately
            
            return grammar_score
        except Exception:
            return 0.8  # Default score on error
    
    def _assess_coherence(self, text: str) -> float:
        """Évalue la cohérence du texte"""
        # Simple coherence assessment based on sentence transitions
        sentences = re.split(r'[.!?]+', text)
        sentences = [s.strip() for s in sentences if s.strip()]
        
        if len(sentences) < 2:
            return 1.0
        
        # Look for transition words
        transition_words = [
            'however', 'therefore', 'furthermore', 'moreover', 'additionally',
            'consequently', 'meanwhile', 'subsequently', 'nevertheless'
        ]
        
        transition_count = sum(1 for sentence in sentences 
                             for word in transition_words 
                             if word in sentence.lower())
        
        coherence_score = min(transition_count / (len(sentences) / 5), 1.0)
        return coherence_score
    
    def _assess_completeness(self, text: str) -> float:
        """Évalue la complétude du contenu"""
        # Basic completeness assessment
        word_count = len(text.split())
        
        # Check for conclusion indicators
        conclusion_indicators = ['conclusion', 'summary', 'in summary', 'to conclude', 'finally']
        has_conclusion = any(indicator in text.lower() for indicator in conclusion_indicators)
        
        # Completeness based on length and structure
        length_score = min(word_count / 500, 1.0) * 0.7
        structure_score = 0.3 if has_conclusion else 0.1
        
        return length_score + structure_score
    
    def _assess_originality(self, text: str) -> float:
        """Évalue l'originalité du contenu"""
        # Basic originality assessment
        # In production, this would check against databases of existing content
        return 0.85  # Placeholder score
    
    def _assess_engagement_potential(self, text: str) -> float:
        """Évalue le potentiel d'engagement"""
        # Look for engaging elements
        engaging_elements = [
            r'\?',  # Questions
            r'!',   # Exclamations
            r'\byou\b',  # Direct address
            r'\bwe\b',   # Inclusive language
            r'\b(amazing|incredible|fantastic|awesome)\b'  # Engaging adjectives
        ]
        
        engagement_count = sum(len(re.findall(pattern, text, re.IGNORECASE)) 
                             for pattern in engaging_elements)
        
        word_count = len(text.split())
        engagement_score = min(engagement_count / (word_count / 100), 1.0)
        
        return engagement_score
    
    def _get_quality_rating(self, score: float) -> str:
        """Convertit le score en rating de qualité"""
        if score >= 0.9:
            return "excellent"
        elif score >= 0.7:
            return "good"
        elif score >= 0.5:
            return "acceptable"
        elif score >= 0.3:
            return "poor"
        else:
            return "very_poor"
    
    def _generate_improvement_suggestions(self, quality_scores: Dict, text_analysis: Dict) -> List[str]:
        """Génère des suggestions d'amélioration"""
        suggestions = []
        
        if quality_scores["grammar_score"] < 0.7:
            suggestions.append("Improve grammar and spelling")
        
        if quality_scores["coherence_score"] < 0.5:
            suggestions.append("Add transition words to improve flow")
        
        if quality_scores["completeness_score"] < 0.6:
            suggestions.append("Add more content or a conclusion")
        
        if quality_scores["engagement_score"] < 0.4:
            suggestions.append("Add questions or direct reader engagement")
        
        return suggestions
    
    def _assess_formality(self, text: str) -> str:
        """Évalue le niveau de formalité"""
        informal_indicators = ['gonna', 'wanna', 'kinda', 'sorta', "don't", "can't", "won't"]
        formal_indicators = ['therefore', 'furthermore', 'consequently', 'nevertheless']
        
        informal_count = sum(text.lower().count(word) for word in informal_indicators)
        formal_count = sum(text.lower().count(word) for word in formal_indicators)
        
        if formal_count > informal_count:
            return "formal"
        elif informal_count > formal_count:
            return "informal"
        else:
            return "neutral"
    
    def _assess_tone(self, text: str) -> str:
        """Évalue le ton du texte"""
        positive_words = ['excellent', 'amazing', 'great', 'wonderful', 'fantastic']
        negative_words = ['terrible', 'awful', 'horrible', 'bad', 'disappointing']
        
        positive_count = sum(text.lower().count(word) for word in positive_words)
        negative_count = sum(text.lower().count(word) for word in negative_words)
        
        if positive_count > negative_count:
            return "positive"
        elif negative_count > positive_count:
            return "negative"
        else:
            return "neutral"
    
    def _generate_heading_recommendations(self, h1: int, h2: int, h3: int) -> List[str]:
        """Génère des recommandations pour les titres"""
        recommendations = []
        
        if h1 == 0:
            recommendations.append("Add a main heading (H1)")
        elif h1 > 1:
            recommendations.append("Use only one H1 tag per document")
        
        if h2 == 0:
            recommendations.append("Add subheadings (H2) to break up content")
        
        if h2 > 0 and h3 == 0:
            recommendations.append("Consider adding H3 tags for better structure")
        
        return recommendations


class AsyncDocumentProcessor(AsyncBaseProcessor):
    """Version asynchrone du processeur de documents"""
    
    def __init__(self, config: Dict[str, Any] = None):
        super().__init__(config)
        self.sync_processor = DocumentProcessor(config)
        self.executor = ThreadPoolExecutor(max_workers=4)
    
    async def validate_input(self, input_data: Any) -> bool:
        """Version asynchrone de la validation"""
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(
            self.executor, 
            self.sync_processor.validate_input, 
            input_data
        )
    
    async def process(self, input_data: Any) -> Dict[str, Any]:
        """Version asynchrone du traitement"""
        loop = asyncio.get_event_loop()
        return await loop.run_in_executor(
            self.executor, 
            self.sync_processor.process, 
            input_data
        )
    
    async def process_batch(self, input_batch: List[Any]) -> List[Dict[str, Any]]:
        """Traitement en lot asynchrone"""
        tasks = [self.process(item) for item in input_batch]
        return await asyncio.gather(*tasks, return_exceptions=True)
