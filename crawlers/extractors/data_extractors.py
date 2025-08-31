"""
Data Extractors - Industrial IA Structured Data Processing System
================================================================

Ultra-advanced professional structured data extractors for databases, APIs, and formatted content.
Implements enterprise-grade data parsing, validation, and transformation capabilities with AI.

Author: Fahed Mlaiel <mlaiel@live.de>
Copyright: All rights reserved. Unauthorized use, reproduction, or distribution prohibited.

WARNING: This code is protected by copyright law. Any unauthorized copying, 
distribution, or modification is strictly prohibited and will result in 
legal action. Contact mlaiel@live.de for licensing.

 STRICT COPYRIGHT PROTECTION 
This code is the intellectual property of Fahed Mlaiel (mlaiel@live.de).
UNAUTHORIZED USE STRICTLY PROHIBITED - Legal action will be taken.

Technical Team Expertise:
- Lead IA Developer: Advanced AI/ML algorithms and neural networks
- Backend Senior: Enterprise architecture and microservices
- ML Engineer: Machine learning pipelines and model optimization
- Database Administrator: Data architecture and optimization
- Security Specialist: Cybersecurity and data protection
- Microservices Architect: Distributed systems and scalability
- Audio Engineer: Digital signal processing and audio analysis
- DevOps Engineer: Infrastructure automation and deployment
- IA Prompt Engineer: Prompt optimization and AI interaction

Project Owner: Fahed Mlaiel - mlaiel@live.de
"""

import asyncio
import logging
import json
import xml.etree.ElementTree as ET
import csv
import io
import re
import time
from typing import Dict, List, Optional, Any, Union, Tuple, Set, Iterator
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from abc import ABC, abstractmethod
from pathlib import Path
from urllib.parse import urlparse, parse_qs
import mimetypes
import hashlib

# Import core extraction components
from .extraction_engine import BaseExtractor, ExtractionRequest, ExtractionResult, ExtractionStatus, ContentType

# Import third-party libraries conditionally
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False

try:
    import openpyxl
    import xlrd
    HAS_EXCEL = True
except ImportError:
    HAS_EXCEL = False

try:
    from bs4 import BeautifulSoup
    HAS_BS4 = True
except ImportError:
    HAS_BS4 = False

try:
    import yaml
    HAS_YAML = True
except ImportError:
    HAS_YAML = False

try:
    import sqlalchemy
    from sqlalchemy import create_engine, text
    HAS_SQLALCHEMY = True
except ImportError:
    HAS_SQLALCHEMY = False

try:
    import pymongo
    HAS_MONGODB = True
except ImportError:
    HAS_MONGODB = False

try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False

logger = logging.getLogger(__name__)


@dataclass
class StructuredDataMetadata:
    """Structured data metadata container"""
    
    format_type: Optional[str] = None
    schema_version: Optional[str] = None
    record_count: int = 0
    column_count: int = 0
    data_types: Dict[str, str] = field(default_factory=dict)
    encoding: Optional[str] = None
    delimiter: Optional[str] = None
    has_header: bool = False
    file_size: int = 0
    validation_errors: List[str] = field(default_factory=list)
    quality_score: float = 0.0
    extracted_schema: Dict[str, Any] = field(default_factory=dict)


class BaseDataExtractor(BaseExtractor):
    """Base class for structured data extractors"""
    
    def __init__(self, name: str, supported_formats: Set[str]):
        super().__init__(name)
        self.supported_formats = supported_formats
        self.max_preview_rows = 1000
        self.chunk_size = 10000
        
    async def validate_data_format(self, content: bytes, format_hint: Optional[str] = None) -> bool:
        """Validate if content matches expected format"""



        return True
    
    async def extract_schema(self, data: Any) -> Dict[str, Any]:
        """Extract data schema information"""



        return {}
    
    async def validate_data_quality(self, data: Any) -> Tuple[float, List[str]]:
        """Validate data quality and return score with errors"""



        return 1.0, []


class JSONExtractor(BaseDataExtractor):
    """Advanced JSON data extractor"""
    
    def __init__(self):
        super().__init__("JSONExtractor", {'.json', '.jsonl', '.ndjson'})
        
    async def can_handle(self, request: ExtractionRequest) -> bool:
        """Check if request contains JSON data"""
        if request.source_path:
            return Path(request.source_path).suffix.lower() in self.supported_formats
        
        if request.source_data:
            try:
                content = request.source_data.decode('utf-8')
                json.loads(content.strip())
                return True
            except (json.JSONDecodeError, UnicodeDecodeError):
                pass
        
        return False
    
    async def extract(self, request: ExtractionRequest) -> ExtractionResult:
        """Extract JSON data and metadata"""



        try:
            # Get content
            if request.source_data:
                content = request.source_data.decode('utf-8')
            elif request.source_path:
                with open(request.source_path, 'r', encoding='utf-8') as f:
                    content = f.read()
            else:
                return ExtractionResult(
                    request_id=request.request_id,
                    status=ExtractionStatus.FAILED,
                    error="No data source provided"
                )
            
            # Parse JSON
            extracted_data = {}
            if content.strip().startswith('[') or content.strip().startswith('{'):
                # Standard JSON
                data = json.loads(content)
                extracted_data = await self._process_json(data)
            else:
                # JSON Lines format
                lines = content.strip().split('\n')
                data = []
                for line in lines:
                    if line.strip():
                        try:
                            data.append(json.loads(line))
                        except json.JSONDecodeError:
                            continue
                extracted_data = await self._process_json_lines(data)
            
            # Extract metadata
            metadata = await self._extract_json_metadata(data, content)
            
            return ExtractionResult(
                request_id=request.request_id,
                status=ExtractionStatus.COMPLETED,
                extracted_data=extracted_data,
                metadata={"structured": metadata},
                content_type=ContentType.STRUCTURED,
                processing_time=time.time() - time.time()
            )
            
        except Exception as e:
            self.logger.error(f"JSON extraction failed: {str(e)}")
            return ExtractionResult(
                request_id=request.request_id,
                status=ExtractionStatus.FAILED,
                error=str(e)
            )
    
    async def _process_json(self, data: Any) -> Dict[str, Any]:
        """Process standard JSON data"""
        result = {
            'type': 'json',
            'data': data,
            'structure': await self._analyze_json_structure(data),
            'statistics': await self._calculate_json_stats(data)
        }
        
        # Extract specific data types
        if isinstance(data, dict):
            result['keys'] = list(data.keys())
            result['nested_levels'] = self._count_nested_levels(data)
        elif isinstance(data, list):
            result['array_length'] = len(data)
            if data:
                result['item_type'] = type(data[0]).__name__
                if isinstance(data[0], dict):
                    result['common_keys'] = self._find_common_keys(data)
        
        return result
    
    async def _process_json_lines(self, data: List[Dict]) -> Dict[str, Any]:
        """Process JSON Lines data"""
        result = {
            'type': 'jsonl',
            'records': data,
            'record_count': len(data),
            'structure': await self._analyze_json_structure(data),
            'statistics': await self._calculate_json_stats(data)
        }
        
        if data:
            result['common_keys'] = self._find_common_keys(data)
            result['key_frequency'] = self._calculate_key_frequency(data)
        
        return result
    
    def _analyze_json_structure(self, data: Any, max_depth: int = 10) -> Dict[str, Any]:
        """Analyze JSON structure recursively"""
        if max_depth <= 0:
            return {'type': 'max_depth_reached'}
        
        if isinstance(data, dict):
            structure = {
                'type': 'object',
                'keys': {},
                'key_count': len(data)
            }
            
            for key, value in data.items():
                structure['keys'][key] = self._analyze_json_structure(value, max_depth - 1)
            
            return structure
        
        elif isinstance(data, list):
            structure = {
                'type': 'array',
                'length': len(data),
                'item_types': {}
            }
            
            # Analyze item types
            for item in data[:100]:  # Sample first 100 items
                item_type = type(item).__name__
                if item_type not in structure['item_types']:
                    structure['item_types'][item_type] = {
                        'count': 0,
                        'structure': self._analyze_json_structure(item, max_depth - 1)
                    }
                structure['item_types'][item_type]['count'] += 1
            
            return structure
        
        else:
            return {
                'type': 'primitive',
                'data_type': type(data).__name__,
                'value': str(data)[:100] if isinstance(data, str) else data
            }
    
    async def _calculate_json_stats(self, data: Any) -> Dict[str, Any]:
        """Calculate JSON statistics"""
        stats = {
            'total_keys': 0,
            'total_values': 0,
            'null_values': 0,
            'empty_strings': 0,
            'nested_objects': 0,
            'arrays': 0,
            'max_depth': 0
        }
        
        def count_recursive(obj, depth=0):
            stats['max_depth'] = max(stats['max_depth'], depth)
            
            if isinstance(obj, dict):
                stats['nested_objects'] += 1
                for key, value in obj.items():
                    stats['total_keys'] += 1
                    stats['total_values'] += 1
                    
                    if value is None:
                        stats['null_values'] += 1
                    elif value == "":
                        stats['empty_strings'] += 1
                    
                    count_recursive(value, depth + 1)
            
            elif isinstance(obj, list):
                stats['arrays'] += 1
                for item in obj:
                    stats['total_values'] += 1
                    count_recursive(item, depth + 1)
        
        count_recursive(data)
        return stats
    
    def _count_nested_levels(self, data: Any, current_level: int = 0) -> int:
        """Count maximum nested levels in JSON"""
        if isinstance(data, dict):
            if not data:
                return current_level
            return max(self._count_nested_levels(value, current_level + 1) 
                      for value in data.values())
        elif isinstance(data, list):
            if not data:
                return current_level
            return max(self._count_nested_levels(item, current_level + 1) 
                      for item in data)
        else:
            return current_level
    
    def _find_common_keys(self, data: List[Dict]) -> List[str]:
        """Find keys common to all objects in array"""
        if not data:
            return []
        
        common_keys = set(data[0].keys()) if isinstance(data[0], dict) else set()
        
        for item in data[1:]:
            if isinstance(item, dict):
                common_keys &= set(item.keys())
            else:
                common_keys = set()
                break
        
        return sorted(list(common_keys))
    
    def _calculate_key_frequency(self, data: List[Dict]) -> Dict[str, int]:
        """Calculate frequency of keys across all objects"""
        frequency = {}
        
        for item in data:
            if isinstance(item, dict):
                for key in item.keys():
                    frequency[key] = frequency.get(key, 0) + 1
        
        return frequency
    
    async def _extract_json_metadata(self, data: Any, content: str) -> StructuredDataMetadata:
        """Extract JSON-specific metadata"""
        stats = await self._calculate_json_stats(data)
        
        return StructuredDataMetadata(
            format_type='json',
            record_count=len(data) if isinstance(data, list) else 1,
            data_types=self._extract_data_types(data),
            encoding='utf-8',
            file_size=len(content.encode('utf-8')),
            quality_score=await self._calculate_json_quality(data),
            extracted_schema=await self.extract_schema(data)
        )
    
    def _extract_data_types(self, data: Any) -> Dict[str, str]:
        """Extract data types from JSON structure"""
        types = {}
        
        def extract_types(obj, path="root"):
            if isinstance(obj, dict):
                for key, value in obj.items():
                    current_path = f"{path}.{key}"
                    types[current_path] = type(value).__name__
                    if isinstance(value, (dict, list)):
                        extract_types(value, current_path)
            elif isinstance(obj, list) and obj:
                types[f"{path}[]"] = f"array of {type(obj[0]).__name__}"
                if isinstance(obj[0], (dict, list)):
                    extract_types(obj[0], f"{path}[0]")
        
        extract_types(data)
        return types
    
    async def _calculate_json_quality(self, data: Any) -> float:
        """Calculate JSON data quality score"""
        score = 1.0
        
        # Check for common quality issues
        stats = await self._calculate_json_stats(data)
        
        if stats['total_values'] > 0:
            null_ratio = stats['null_values'] / stats['total_values']
            empty_ratio = stats['empty_strings'] / stats['total_values']
            
            # Deduct points for missing data
            score -= null_ratio * 0.3
            score -= empty_ratio * 0.2
        
        # Check for consistent structure in arrays
        if isinstance(data, list) and data:
            consistent_keys = self._check_array_consistency(data)
            if not consistent_keys:
                score -= 0.3
        
        return max(0.0, score)
    
    def _check_array_consistency(self, data: List) -> bool:
        """Check if array objects have consistent structure"""
        if not data or not isinstance(data[0], dict):
            return True
        
        reference_keys = set(data[0].keys())
        
        for item in data[1:]:
            if not isinstance(item, dict) or set(item.keys()) != reference_keys:
                return False
        
        return True


class CSVExtractor(BaseDataExtractor):
    """Advanced CSV data extractor"""
    
    def __init__(self):
        super().__init__("CSVExtractor", {'.csv', '.tsv', '.tab'})
        
    async def can_handle(self, request: ExtractionRequest) -> bool:
        """Check if request contains CSV data"""
        if request.source_path:
            return Path(request.source_path).suffix.lower() in self.supported_formats
        
        if request.source_data:
            try:
                content = request.source_data.decode('utf-8')
                # Try to detect CSV structure
                return self._detect_csv_structure(content) is not None
            except UnicodeDecodeError:
                return False
        
        return False
    
    async def extract(self, request: ExtractionRequest) -> ExtractionResult:
        """Extract CSV data and metadata"""



        try:
            # Get content
            if request.source_data:
                content = request.source_data.decode('utf-8')
            elif request.source_path:
                with open(request.source_path, 'r', encoding='utf-8') as f:
                    content = f.read()
            else:
                return ExtractionResult(
                    request_id=request.request_id,
                    status=ExtractionStatus.FAILED,
                    error="No data source provided"
                )
            
            # Detect CSV structure
            csv_config = self._detect_csv_structure(content)
            if not csv_config:
                return ExtractionResult(
                    request_id=request.request_id,
                    status=ExtractionStatus.FAILED,
                    error="Cannot detect CSV structure"
                )
            
            # Parse CSV
            extracted_data = await self._process_csv(content, csv_config)
            
            # Extract metadata
            metadata = await self._extract_csv_metadata(extracted_data, content, csv_config)
            
            return ExtractionResult(
                request_id=request.request_id,
                status=ExtractionStatus.COMPLETED,
                extracted_data=extracted_data,
                metadata={"structured": metadata},
                content_type=ContentType.STRUCTURED,
                processing_time=time.time() - time.time()
            )
            
        except Exception as e:
            self.logger.error(f"CSV extraction failed: {str(e)}")
            return ExtractionResult(
                request_id=request.request_id,
                status=ExtractionStatus.FAILED,
                error=str(e)
            )
    
    def _detect_csv_structure(self, content: str) -> Optional[Dict[str, Any]]:
        """Detect CSV delimiter, quoting, and structure"""
        # Try common delimiters
        delimiters = [',', ';', '\t', '|']
        
        lines = content.strip().split('\n')[:10]  # Sample first 10 lines
        
        best_config = None
        best_score = 0
        
        for delimiter in delimiters:
            try:
                reader = csv.reader(lines, delimiter=delimiter)
                rows = list(reader)
                
                if len(rows) < 2:
                    continue
                
                # Calculate consistency score
                header_len = len(rows[0])
                consistency_score = sum(1 for row in rows[1:] if len(row) == header_len)
                consistency_score /= len(rows) - 1
                
                if consistency_score > best_score:
                    best_score = consistency_score
                    best_config = {
                        'delimiter': delimiter,
                        'has_header': self._detect_header(rows),
                        'column_count': header_len,
                        'quoting': csv.QUOTE_MINIMAL
                    }
            except Exception:
                continue
        
        return best_config if best_score > 0.7 else None
    
    def _detect_header(self, rows: List[List[str]]) -> bool:
        """Detect if first row is header"""
        if len(rows) < 2:
            return False
        
        header_row = rows[0]
        data_rows = rows[1:]
        
        # Check if header contains non-numeric strings while data contains numbers
        header_numeric = sum(1 for cell in header_row if self._is_numeric(cell))
        data_numeric = sum(1 for row in data_rows for cell in row if self._is_numeric(cell))
        
        header_numeric_ratio = header_numeric / len(header_row) if header_row else 0
        data_numeric_ratio = data_numeric / (len(data_rows) * len(header_row)) if data_rows and header_row else 0
        
        # Header likely if it has fewer numbers than data
        return header_numeric_ratio < data_numeric_ratio * 0.5
    
    def _is_numeric(self, value: str) -> bool:
        """Check if string represents a number"""



        try:
            float(value)
            return True
        except (ValueError, TypeError):
            return False
    
    async def _process_csv(self, content: str, config: Dict[str, Any]) -> Dict[str, Any]:
        """Process CSV data"""
        reader = csv.reader(
            io.StringIO(content),
            delimiter=config['delimiter'],
            quoting=config['quoting']
        )
        
        rows = list(reader)
        
        result = {
            'type': 'csv',
            'delimiter': config['delimiter'],
            'has_header': config['has_header'],
            'total_rows': len(rows),
            'column_count': config['column_count']
        }
        
        if config['has_header'] and rows:
            result['headers'] = rows[0]
            result['data'] = rows[1:]
            result['data_rows'] = len(rows) - 1
        else:
            result['headers'] = [f"column_{i}" for i in range(config['column_count'])]
            result['data'] = rows
            result['data_rows'] = len(rows)
        
        # Analyze columns
        if result['data']:
            result['column_analysis'] = await self._analyze_csv_columns(
                result['data'], result['headers']
            )
        
        # Sample data for preview
        result['preview'] = result['data'][:self.max_preview_rows]
        
        return result
    
    async def _analyze_csv_columns(self, data: List[List[str]], headers: List[str]) -> Dict[str, Any]:
        """Analyze CSV columns for data types and statistics"""
        analysis = {}
        
        for col_idx, header in enumerate(headers):
            column_data = [row[col_idx] if col_idx < len(row) else '' for row in data]
            
            analysis[header] = {
                'index': col_idx,
                'data_type': self._detect_column_type(column_data),
                'null_count': sum(1 for cell in column_data if not cell.strip()),
                'unique_count': len(set(column_data)),
                'max_length': max(len(cell) for cell in column_data) if column_data else 0,
                'sample_values': list(set(column_data))[:10]
            }
            
            # Numeric statistics
            numeric_values = [float(cell) for cell in column_data 
                            if self._is_numeric(cell) and cell.strip()]
            
            if numeric_values:
                analysis[header]['numeric_stats'] = {
                    'min': min(numeric_values),
                    'max': max(numeric_values),
                    'mean': sum(numeric_values) / len(numeric_values),
                    'count': len(numeric_values)
                }
        
        return analysis
    
    def _detect_column_type(self, column_data: List[str]) -> str:
        """Detect column data type"""
        non_empty = [cell.strip() for cell in column_data if cell.strip()]
        
        if not non_empty:
            return 'empty'
        
        # Check for numeric
        numeric_count = sum(1 for cell in non_empty if self._is_numeric(cell))
        numeric_ratio = numeric_count / len(non_empty)
        
        if numeric_ratio > 0.8:
            # Check if integers
            integer_count = sum(1 for cell in non_empty 
                              if self._is_numeric(cell) and float(cell).is_integer())
            if integer_count / len(non_empty) > 0.8:
                return 'integer'
            else:
                return 'float'
        
        # Check for dates
        date_count = sum(1 for cell in non_empty if self._is_date(cell))
        if date_count / len(non_empty) > 0.8:
            return 'date'
        
        # Check for booleans
        boolean_values = {'true', 'false', '1', '0', 'yes', 'no', 'y', 'n'}
        boolean_count = sum(1 for cell in non_empty 
                          if cell.lower() in boolean_values)
        if boolean_count / len(non_empty) > 0.8:
            return 'boolean'
        
        return 'string'
    
    def _is_date(self, value: str) -> bool:
        """Check if string represents a date"""
        date_patterns = [
            r'\d{4}-\d{2}-\d{2}',  # YYYY-MM-DD
            r'\d{2}/\d{2}/\d{4}',  # MM/DD/YYYY
            r'\d{2}-\d{2}-\d{4}',  # MM-DD-YYYY
            r'\d{4}/\d{2}/\d{2}',  # YYYY/MM/DD
        ]
        
        return any(re.match(pattern, value.strip()) for pattern in date_patterns)
    
    async def _extract_csv_metadata(self, data: Dict, content: str, config: Dict) -> StructuredDataMetadata:
        """Extract CSV-specific metadata"""



        return StructuredDataMetadata(
            format_type='csv',
            record_count=data.get('data_rows', 0),
            column_count=data.get('column_count', 0),
            delimiter=config['delimiter'],
            has_header=config['has_header'],
            encoding='utf-8',
            file_size=len(content.encode('utf-8')),
            quality_score=await self._calculate_csv_quality(data),
            extracted_schema=await self.extract_schema(data)
        )
    
    async def _calculate_csv_quality(self, data: Dict) -> float:
        """Calculate CSV data quality score"""
        score = 1.0
        
        column_analysis = data.get('column_analysis', {})
        total_cells = data.get('data_rows', 0) * data.get('column_count', 1)
        
        if total_cells > 0:
            # Calculate null ratio
            total_nulls = sum(col.get('null_count', 0) for col in column_analysis.values())
            null_ratio = total_nulls / total_cells
            score -= null_ratio * 0.4
            
            # Check for consistent data types
            type_consistency = sum(1 for col in column_analysis.values() 
                                 if col.get('data_type') not in ['empty', 'string'])
            type_consistency /= len(column_analysis) if column_analysis else 1
            score *= type_consistency
        
        return max(0.0, score)


class XMLExtractor(BaseDataExtractor):
    """Advanced XML data extractor"""
    
    def __init__(self):
        super().__init__("XMLExtractor", {'.xml', '.xsl', '.xsd', '.rss', '.atom'})
        
    async def can_handle(self, request: ExtractionRequest) -> bool:
        """Check if request contains XML data"""
        if request.source_path:
            return Path(request.source_path).suffix.lower() in self.supported_formats
        
        if request.source_data:
            try:
                content = request.source_data.decode('utf-8').strip()
                return content.startswith('<?xml') or content.startswith('<')
            except UnicodeDecodeError:
                return False
        
        return False
    
    async def extract(self, request: ExtractionRequest) -> ExtractionResult:
        """Extract XML data and metadata"""



        try:
            # Get content
            if request.source_data:
                content = request.source_data.decode('utf-8')
            elif request.source_path:
                with open(request.source_path, 'r', encoding='utf-8') as f:
                    content = f.read()
            else:
                return ExtractionResult(
                    request_id=request.request_id,
                    status=ExtractionStatus.FAILED,
                    error="No data source provided"
                )
            
            # Parse XML
            try:
                root = ET.fromstring(content)
            except ET.ParseError as e:
                return ExtractionResult(
                    request_id=request.request_id,
                    status=ExtractionStatus.FAILED,
                    error=f"XML parsing error: {str(e)}"
                )
            
            # Process XML
            extracted_data = await self._process_xml(root, content)
            
            # Extract metadata
            metadata = await self._extract_xml_metadata(extracted_data, content)
            
            return ExtractionResult(
                request_id=request.request_id,
                status=ExtractionStatus.COMPLETED,
                extracted_data=extracted_data,
                metadata={"structured": metadata},
                content_type=ContentType.STRUCTURED,
                processing_time=time.time() - time.time()
            )
            
        except Exception as e:
            self.logger.error(f"XML extraction failed: {str(e)}")
            return ExtractionResult(
                request_id=request.request_id,
                status=ExtractionStatus.FAILED,
                error=str(e)
            )
    
    async def _process_xml(self, root: ET.Element, content: str) -> Dict[str, Any]:
        """Process XML data"""
        result = {
            'type': 'xml',
            'root_tag': root.tag,
            'namespaces': self._extract_namespaces(content),
            'structure': await self._analyze_xml_structure(root),
            'statistics': await self._calculate_xml_stats(root)
        }
        
        # Convert to dictionary
        result['data'] = self._xml_to_dict(root)
        
        # Extract specific XML types
        if 'rss' in root.tag.lower() or 'feed' in root.tag.lower():
            result['feed_type'] = 'rss' if 'rss' in root.tag.lower() else 'atom'
            result['feed_data'] = await self._extract_feed_data(root)
        
        return result
    
    def _extract_namespaces(self, content: str) -> Dict[str, str]:
        """Extract XML namespaces"""
        namespaces = {}
        namespace_pattern = r'xmlns(?::([^=]+))?="([^"]+)"'
        
        for match in re.finditer(namespace_pattern, content):
            prefix = match.group(1) or 'default'
            uri = match.group(2)
            namespaces[prefix] = uri
        
        return namespaces
    
    async def _analyze_xml_structure(self, element: ET.Element, max_depth: int = 10) -> Dict[str, Any]:
        """Analyze XML structure recursively"""
        if max_depth <= 0:
            return {'type': 'max_depth_reached'}
        
        structure = {
            'tag': element.tag,
            'attributes': dict(element.attrib),
            'has_text': bool(element.text and element.text.strip()),
            'children': {}
        }
        
        # Analyze child elements
        child_tags = {}
        for child in element:
            tag = child.tag
            if tag not in child_tags:
                child_tags[tag] = {
                    'count': 0,
                    'structure': self._analyze_xml_structure(child, max_depth - 1)
                }
            child_tags[tag]['count'] += 1
        
        structure['children'] = child_tags
        structure['child_count'] = len(list(element))
        
        return structure
    
    async def _calculate_xml_stats(self, root: ET.Element) -> Dict[str, Any]:
        """Calculate XML statistics"""
        stats = {
            'total_elements': 0,
            'total_attributes': 0,
            'max_depth': 0,
            'text_elements': 0,
            'empty_elements': 0,
            'unique_tags': set()
        }
        
        def count_recursive(element, depth=0):
            stats['total_elements'] += 1
            stats['max_depth'] = max(stats['max_depth'], depth)
            stats['total_attributes'] += len(element.attrib)
            stats['unique_tags'].add(element.tag)
            
            if element.text and element.text.strip():
                stats['text_elements'] += 1
            
            if not list(element) and not (element.text and element.text.strip()):
                stats['empty_elements'] += 1
            
            for child in element:
                count_recursive(child, depth + 1)
        
        count_recursive(root)
        stats['unique_tag_count'] = len(stats['unique_tags'])
        del stats['unique_tags']  # Convert set to count
        
        return stats
    
    def _xml_to_dict(self, element: ET.Element) -> Dict[str, Any]:
        """Convert XML element to dictionary"""
        result = {}
        
        # Add attributes
        if element.attrib:
            result['@attributes'] = dict(element.attrib)
        
        # Add text content
        if element.text and element.text.strip():
            if list(element):
                result['#text'] = element.text.strip()
            else:
                return element.text.strip()
        
        # Add child elements
        children = {}
        for child in element:
            child_dict = self._xml_to_dict(child)
            
            if child.tag in children:
                # Multiple elements with same tag - convert to list
                if not isinstance(children[child.tag], list):
                    children[child.tag] = [children[child.tag]]
                children[child.tag].append(child_dict)
            else:
                children[child.tag] = child_dict
        
        result.update(children)
        return result
    
    async def _extract_feed_data(self, root: ET.Element) -> Dict[str, Any]:
        """Extract RSS/Atom feed specific data"""
        feed_data = {}
        
        if 'rss' in root.tag.lower():
            # RSS feed
            channel = root.find('channel')
            if channel is not None:
                feed_data['title'] = self._get_element_text(channel, 'title')
                feed_data['description'] = self._get_element_text(channel, 'description')
                feed_data['link'] = self._get_element_text(channel, 'link')
                
                items = channel.findall('item')
                feed_data['item_count'] = len(items)
                feed_data['items'] = []
                
                for item in items[:10]:  # Sample first 10 items
                    item_data = {
                        'title': self._get_element_text(item, 'title'),
                        'description': self._get_element_text(item, 'description'),
                        'link': self._get_element_text(item, 'link'),
                        'pubDate': self._get_element_text(item, 'pubDate')
                    }
                    feed_data['items'].append(item_data)
        
        elif 'feed' in root.tag.lower():
            # Atom feed
            feed_data['title'] = self._get_element_text(root, 'title')
            feed_data['subtitle'] = self._get_element_text(root, 'subtitle')
            
            entries = root.findall('.//{http://www.w3.org/2005/Atom}entry')
            feed_data['entry_count'] = len(entries)
            feed_data['entries'] = []
            
            for entry in entries[:10]:  # Sample first 10 entries
                entry_data = {
                    'title': self._get_element_text(entry, '{http://www.w3.org/2005/Atom}title'),
                    'summary': self._get_element_text(entry, '{http://www.w3.org/2005/Atom}summary'),
                    'updated': self._get_element_text(entry, '{http://www.w3.org/2005/Atom}updated')
                }
                feed_data['entries'].append(entry_data)
        
        return feed_data
    
    def _get_element_text(self, parent: ET.Element, tag: str) -> Optional[str]:
        """Get text content of child element"""
        element = parent.find(tag)
        return element.text if element is not None and element.text else None
    
    async def _extract_xml_metadata(self, data: Dict, content: str) -> StructuredDataMetadata:
        """Extract XML-specific metadata"""
        stats = data.get('statistics', {})
        
        return StructuredDataMetadata(
            format_type='xml',
            record_count=stats.get('total_elements', 0),
            encoding='utf-8',
            file_size=len(content.encode('utf-8')),
            quality_score=await self._calculate_xml_quality(data),
            extracted_schema=await self.extract_schema(data)
        )
    
    async def _calculate_xml_quality(self, data: Dict) -> float:
        """Calculate XML data quality score"""
        score = 1.0
        
        stats = data.get('statistics', {})
        
        # Check for empty elements
        total_elements = stats.get('total_elements', 1)
        empty_elements = stats.get('empty_elements', 0)
        empty_ratio = empty_elements / total_elements
        
        score -= empty_ratio * 0.3
        
        # Check for proper structure
        if stats.get('max_depth', 0) < 2:
            score -= 0.2  # Very flat structure might indicate poor design
        
        return max(0.0, score)


class ExcelExtractor(BaseDataExtractor):
    """Advanced Excel data extractor"""
    
    def __init__(self):
        super().__init__("ExcelExtractor", {'.xlsx', '.xls', '.xlsm'})
        
    async def can_handle(self, request: ExtractionRequest) -> bool:
        """Check if request contains Excel data"""
        if not HAS_EXCEL:
            return False
            
        if request.source_path:
            return Path(request.source_path).suffix.lower() in self.supported_formats
        
        return False
    
    async def extract(self, request: ExtractionRequest) -> ExtractionResult:
        """Extract Excel data and metadata"""
        if not HAS_EXCEL:
            return ExtractionResult(
                request_id=request.request_id,
                status=ExtractionStatus.FAILED,
                error="Excel processing libraries not available"
            )
        
        try:
            if not request.source_path:
                return ExtractionResult(
                    request_id=request.request_id,
                    status=ExtractionStatus.FAILED,
                    error="Excel extraction requires file path"
                )
            
            # Process Excel file
            extracted_data = await self._process_excel(request.source_path)
            
            # Extract metadata
            metadata = await self._extract_excel_metadata(extracted_data)
            
            return ExtractionResult(
                request_id=request.request_id,
                status=ExtractionStatus.COMPLETED,
                extracted_data=extracted_data,
                metadata={"structured": metadata},
                content_type=ContentType.STRUCTURED,
                processing_time=time.time() - time.time()
            )
            
        except Exception as e:
            self.logger.error(f"Excel extraction failed: {str(e)}")
            return ExtractionResult(
                request_id=request.request_id,
                status=ExtractionStatus.FAILED,
                error=str(e)
            )
    
    async def _process_excel(self, file_path: str) -> Dict[str, Any]:
        """Process Excel file"""
        if HAS_PANDAS:
            # Use pandas for comprehensive analysis
            excel_file = pd.ExcelFile(file_path)
            
            result = {
                'type': 'excel',
                'file_path': file_path,
                'sheet_names': excel_file.sheet_names,
                'sheet_count': len(excel_file.sheet_names),
                'sheets': {}
            }
            
            for sheet_name in excel_file.sheet_names:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                
                sheet_data = {
                    'name': sheet_name,
                    'shape': df.shape,
                    'columns': df.columns.tolist(),
                    'dtypes': df.dtypes.to_dict(),
                    'null_counts': df.isnull().sum().to_dict(),
                    'sample_data': df.head(10).to_dict('records'),
                    'statistics': {}
                }
                
                # Numeric columns statistics
                numeric_cols = df.select_dtypes(include=[np.number]).columns
                if len(numeric_cols) > 0:
                    sheet_data['statistics']['numeric'] = df[numeric_cols].describe().to_dict()
                
                result['sheets'][sheet_name] = sheet_data
            
            return result
        
        else:
            # Fallback to openpyxl
            from openpyxl import load_workbook
            
            wb = load_workbook(file_path, read_only=True)
            
            result = {
                'type': 'excel',
                'file_path': file_path,
                'sheet_names': wb.sheetnames,
                'sheet_count': len(wb.sheetnames),
                'sheets': {}
            }
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Get dimensions
                max_row = ws.max_row
                max_col = ws.max_column
                
                # Extract sample data
                data = []
                for row in ws.iter_rows(min_row=1, max_row=min(11, max_row), values_only=True):
                    data.append(list(row))
                
                sheet_data = {
                    'name': sheet_name,
                    'shape': (max_row, max_col),
                    'sample_data': data,
                    'has_data': max_row > 0 and max_col > 0
                }
                
                result['sheets'][sheet_name] = sheet_data
            
            return result
    
    async def _extract_excel_metadata(self, data: Dict) -> StructuredDataMetadata:
        """Extract Excel-specific metadata"""
        total_rows = sum(sheet.get('shape', [0, 0])[0] for sheet in data.get('sheets', {}).values())
        total_cols = sum(sheet.get('shape', [0, 0])[1] for sheet in data.get('sheets', {}).values())
        
        return StructuredDataMetadata(
            format_type='excel',
            record_count=total_rows,
            column_count=total_cols,
            encoding='binary',
            file_size=Path(data.get('file_path', '')).stat().st_size if data.get('file_path') else 0,
            quality_score=await self._calculate_excel_quality(data)
        )
    
    async def _calculate_excel_quality(self, data: Dict) -> float:
        """Calculate Excel data quality score"""
        score = 1.0
        
        sheets = data.get('sheets', {})
        if not sheets:
            return 0.0
        
        for sheet_data in sheets.values():
            if HAS_PANDAS and 'null_counts' in sheet_data:
                # Calculate null ratio
                null_counts = sheet_data['null_counts']
                total_cells = sheet_data['shape'][0] * sheet_data['shape'][1]
                null_ratio = sum(null_counts.values()) / total_cells if total_cells > 0 else 0
                score -= null_ratio * 0.3
        
        return max(0.0, score / len(sheets))


# Data Extractor Factory
class DataExtractorFactory:
    """Factory for creating data extractors"""
    
    _extractors: List[BaseDataExtractor] = []
    
    @classmethod
    def register_extractor(cls, extractor: BaseDataExtractor):
        """Register a data extractor"""
        cls._extractors.append(extractor)
    
    @classmethod
    def get_extractor(cls, request: ExtractionRequest) -> Optional[BaseDataExtractor]:
        """Get appropriate extractor for request"""
        for extractor in cls._extractors:
            if asyncio.run(extractor.can_handle(request)):
                return extractor
        return None
    
    @classmethod
    def list_supported_formats(cls) -> Set[str]:
        """List all supported data formats"""
        formats = set()
        for extractor in cls._extractors:
            formats.update(extractor.supported_formats)
        return formats


# Register default extractors
def register_default_data_extractors():
    """Register all default data extractors"""
    factory = DataExtractorFactory
    
    factory.register_extractor(JSONExtractor())
    factory.register_extractor(CSVExtractor())
    factory.register_extractor(XMLExtractor())
    factory.register_extractor(ExcelExtractor())


# Initialize on import
register_default_data_extractors()


__all__ = [
    'StructuredDataMetadata',
    'BaseDataExtractor',
    'JSONExtractor',
    'CSVExtractor', 
    'XMLExtractor',
    'ExcelExtractor',
    'DataExtractorFactory',
    'register_default_data_extractors'
]
